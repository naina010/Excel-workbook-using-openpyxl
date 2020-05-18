import openpyxl as xl

fileName = "Book1.xlxs"
isOpen = False
heading = []
data = []
book = 0
sheet = 0

def createWorkbook():
	global isOpen
	global book
	global sheet
	global data
	global heading
  
	try:
		if isOpen == False:
			isOpen = True
			book = xl.Workbook()
			sheet = book.active
      
      # input column names
			while True : 
				head = input("Enter Heading (Exit for exit): ")
				if head.find("Exit") > -1 :
					break
				heading.append(head)
        
			print("\nAdd Row with (,) Separation or press Exit for Exit")
			print("\n***************************************************")
			for head in heading:
				print(head,end="\t\t")
			print("\n***************************************************")
      
      # input data row wise
			while True:
				row = input("")
				if row.find("Exit") > -1:
					break
				data.append(row.split(","))			
        
		else:
			choice = input("Do you want to save changes to "+fileName)
			if choice.find("save") > -1 or choice .find("Save"):
				saveWorkbook()

	except Exception as e:
		print("File already in use : ",e)

def openWorkbook():
	try:
		tempName = input("Enter file Name to be Open : ")
		if tempName.find("Exit")!=-1:
			return
		if tempName.find(".xlsx")==-1:
			tempName+=".xlsx"
      
		print("File found")
		sheet = xl.load_workbook(tempName)
		sheet_obj = sheet.active
		row,col = sheet_obj.max_row, sheet_obj.max_column
    
		for i in range(1,row+1):
			for j in range(1,col+1):
				print(sheet_obj.cell(row=i,column=j).value, end = '\t')
			print()
			print('-------------------------------------------------------------')

	except Exception as e:
		print("File not found ",e)
		openWorkbook()

def saveWorkbook():
	global fileName
	global heading
	global data
	global isOpen
	global book
	global sheet
  
	if isOpen == True:
		sheet.append(heading)
		for row in data:
			sheet.append(row)
      
		fileName = input("Enter File Name : ")
		if fileName.find(".xlsx") == -1:
			fileName = fileName + ".xlsx"		
		book.save(fileName)
		print("File saved.")
    
	else:
		print("No file opened")


def main():
	global fileName
	while True:
		print("1. Create new Workbook")
		print("2. Open Workbook")
		print("3. Save Workbook")
		print("0. Exit")
    
		choice = int(input("Enter Choice : "))
		if choice == 1:
			createWorkbook()
		elif choice == 2:
			openWorkbook()
		elif choice == 3:
			saveWorkbook()
		elif choice == 0:
			break
		else:
			print("Invalid Choice")


if __name__ == "__main__":
	main()
